# © Todos os direitos reservados – github.com/Wbad-02
import io
from models import agora as _agora_br
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from auth import get_usuario_atual, requer_admin, requer_editor_ou_admin
import models

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/api/relatorios", tags=["relatorios"])


def _obter_materiais(db: Session, apenas_alertas: bool):
    mats = (
        db.query(models.Material)
        .filter(models.Material.ativo == True)
        .join(models.GrupoMaterial)
        .order_by(models.GrupoMaterial.nome, models.Material.nome)
        .all()
    )
    if apenas_alertas:
        mats = [m for m in mats if m.alerta_minimo]
    return mats


@router.get("/excel")
def exportar_excel(
    apenas_alertas: bool = False,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    materiais = _obter_materiais(db, apenas_alertas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque"

    header_fill = PatternFill("solid", fgColor="1E3A34")
    alert_fill  = PatternFill("solid", fgColor="FFF3CD")
    header_font = Font(bold=True, color="FFFFFF")

    headers    = ["ID", "Material", "Categoria", "Grupo", "Qtd.", "Unidade", "Mín. Grupo", "Status"]
    col_widths = [6, 28, 18, 18, 8, 8, 12, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, m in enumerate(materiais, 2):
        alerta = m.alerta_minimo
        status = "⚠ ALERTA" if alerta else "OK"
        row_data = [
            m.id, m.nome,
            m.categoria.nome if m.categoria else "—",
            m.grupo.nome,
            m.quantidade, m.unidade,
            m.grupo.quantidade_minima,
            status,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")
            if alerta:
                cell.fill = alert_fill

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"estoque_{_agora_br().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/pdf")
def exportar_pdf(
    apenas_alertas: bool = False,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    materiais = _obter_materiais(db, apenas_alertas)

    output = io.BytesIO()
    doc    = SimpleDocTemplate(output, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1E3A34"),
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey,
    )

    elements = []
    titulo = "Relatório de Estoque" + (" – Itens em Alerta" if apenas_alertas else "")
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(
        f"Gerado em {_agora_br().strftime('%d/%m/%Y às %H:%M')}",
        sub_style,
    ))
    elements.append(Spacer(1, 0.5*cm))

    table_data = [["Material", "Categoria", "Grupo", "Qtd.", "Un.", "Mín.", "Status"]]
    for m in materiais:
        alerta = m.alerta_minimo
        table_data.append([
            m.nome,
            m.categoria.nome if m.categoria else "—",
            m.grupo.nome,
            str(m.quantidade),
            m.unidade,
            str(m.grupo.quantidade_minima),
            "ALERTA" if alerta else "OK",
        ])

    col_widths_pdf = [5.5*cm, 3.5*cm, 3.5*cm, 1.8*cm, 1.5*cm, 1.5*cm, 2*cm]
    table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A34")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))

    for row_idx, m in enumerate(materiais, 1):
        if m.alerta_minimo:
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#FFF3CD")),
                ("TEXTCOLOR",  (-1, row_idx), (-1, row_idx), colors.HexColor("#856404")),
            ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(
        f"Total: {len(materiais)} item(ns) · © Todos os direitos reservados – github.com/Wbad-02",
        sub_style,
    ))

    doc.build(elements)
    output.seek(0)

    filename = f"estoque_{_agora_br().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/saidas/excel")
def exportar_saidas_excel(
    motivo:       str | None = None,
    data_inicio:  str | None = None,
    data_fim:     str | None = None,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    from datetime import datetime as dt
    q = db.query(models.Movimentacao).filter(models.Movimentacao.tipo == "saida")
    if motivo:
        q = q.filter(models.Movimentacao.motivo == motivo)
    if data_inicio:
        q = q.filter(models.Movimentacao.criado_em >= dt.fromisoformat(data_inicio))
    if data_fim:
        q = q.filter(models.Movimentacao.criado_em <= dt.fromisoformat(data_fim + "T23:59:59"))
    rows = q.order_by(models.Movimentacao.criado_em.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saídas"

    header_fill = PatternFill("solid", fgColor="1E3A34")
    header_font = Font(bold=True, color="FFFFFF")
    colab_fill  = PatternFill("solid", fgColor="E3F2FD")
    defeito_fill = PatternFill("solid", fgColor="FCE4EC")

    headers    = ["Data/Hora", "Material", "Categoria", "Grupo", "Qtd.", "Unidade", "Motivo", "Observação", "Usuário"]
    col_widths = [18, 28, 18, 18, 8, 8, 18, 30, 18]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    motivo_label = {"colaborador": "Atribuído a colaborador", "defeito": "Defeito / Ruim"}

    for ri, r in enumerate(rows, 2):
        mat  = r.material
        grp  = mat.grupo       if mat  else None
        cat  = grp.categoria   if grp  else None
        usr  = r.usuario
        motivo_str = r.motivo if r.motivo else ""
        row_data = [
            r.criado_em.strftime("%d/%m/%Y %H:%M"),
            mat.nome           if mat  else "—",
            cat.nome           if cat  else "—",
            grp.nome           if grp  else "—",
            r.quantidade,
            mat.unidade        if mat  else "—",
            motivo_label.get(motivo_str, motivo_str or "—"),
            r.observacao or "—",
            usr.nome           if usr  else "Sistema",
        ]
        fill = colab_fill if motivo_str == "colaborador" else defeito_fill
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center" if ci not in (2,8) else "left")
            cell.fill = fill

    ws.freeze_panes = "A2"

    # Totais no rodapé
    ws.cell(row=len(rows)+3, column=1, value="Total de registros:")
    ws.cell(row=len(rows)+3, column=2, value=len(rows)).font = Font(bold=True)
    ws.cell(row=len(rows)+4, column=1, value="© Todos os direitos reservados – github.com/Wbad-02")

    output = io.BytesIO(); wb.save(output); output.seek(0)
    filename = f"saidas_{_agora_br().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/saidas/pdf")
def exportar_saidas_pdf(
    motivo:      str | None = None,
    data_inicio: str | None = None,
    data_fim:    str | None = None,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    from datetime import datetime as dt
    q = db.query(models.Movimentacao).filter(models.Movimentacao.tipo == "saida")
    if motivo:
        q = q.filter(models.Movimentacao.motivo == motivo)
    if data_inicio:
        q = q.filter(models.Movimentacao.criado_em >= dt.fromisoformat(data_inicio))
    if data_fim:
        q = q.filter(models.Movimentacao.criado_em <= dt.fromisoformat(data_fim + "T23:59:59"))
    rows = q.order_by(models.Movimentacao.criado_em.desc()).all()

    output = io.BytesIO()
    doc    = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=14, textColor=colors.HexColor("#1E3A34"))
    sub_style   = ParagraphStyle("s", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements = []
    titulo = "Relatório de Saídas de Estoque"
    if motivo:
        _motivo_label = {'colaborador': 'Colaborador', 'defeito': 'Defeito'}
        titulo += f" — {_motivo_label.get(motivo, motivo)}"
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(f"Gerado em {_agora_br().strftime('%d/%m/%Y às %H:%M')}", sub_style))
    elements.append(Spacer(1, 0.4*cm))

    motivo_label = {"colaborador": "Colaborador", "defeito": "Defeito"}
    table_data = [["Data/Hora", "Material", "Grupo", "Qtd.", "Motivo", "Observação", "Usuário"]]
    for r in rows:
        mat = r.material; grp = mat.grupo if mat else None; usr = r.usuario
        motivo_str = r.motivo if r.motivo else ""
        table_data.append([
            r.criado_em.strftime("%d/%m %H:%M"),
            mat.nome        if mat else "—",
            grp.nome        if grp else "—",
            str(r.quantidade),
            motivo_label.get(motivo_str, motivo_str or "—"),
            (r.observacao or "—")[:30],
            usr.nome        if usr else "Sistema",
        ])

    cw = [3*cm, 5*cm, 3*cm, 1.5*cm, 2.5*cm, 3.5*cm, 3*cm]
    table = Table(table_data, colWidths=cw, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E3A34")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(
        f"Total: {len(rows)} registro(s) · © Todos os direitos reservados – github.com/Wbad-02",
        sub_style,
    ))
    doc.build(elements)
    output.seek(0)

    filename = f"saidas_{_agora_br().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Relatório de Ativos ───────────────────────────────────────

@router.get("/ativos/excel")
def exportar_ativos_excel(
    status: str | None = None,   # "ativo" | "inativo" | None = todos
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    q = db.query(models.Ativo)
    if status == "ativo":
        q = q.filter(models.Ativo.ativo == True)
    elif status == "inativo":
        q = q.filter(models.Ativo.ativo == False)
    ativos = q.order_by(models.Ativo.nome).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ativos"
    header_fill = PatternFill("solid", fgColor="1E3A34")
    header_font = Font(bold=True, color="FFFFFF")
    inativo_fill = PatternFill("solid", fgColor="ECEFF1")

    headers    = ["ID", "Nome", "Descrição", "Categoria", "Grupo", "Status", "Materiais em uso", "Cadastrado em"]
    col_widths = [6, 28, 22, 18, 18, 10, 16, 18]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for ri, a in enumerate(ativos, 2):
        itens_ativos = sum(1 for i in a.itens if i.devolvido_em is None)
        row_data = [
            a.id, a.nome, a.descricao or "—",
            a.grupo.categoria.nome if a.grupo and a.grupo.categoria else "—",
            a.grupo.nome if a.grupo else "—",
            "Ativo" if a.ativo else "Inativo",
            itens_ativos,
            a.criado_em.strftime("%d/%m/%Y"),
        ]
        fill = inativo_fill if not a.ativo else None
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center" if ci != 2 else "left")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    output = io.BytesIO(); wb.save(output); output.seek(0)
    filename = f"ativos_{_agora_br().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/ativos/pdf")
def exportar_ativos_pdf(
    status: str | None = None,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    q = db.query(models.Ativo)
    if status == "ativo":
        q = q.filter(models.Ativo.ativo == True)
    elif status == "inativo":
        q = q.filter(models.Ativo.ativo == False)
    ativos = q.order_by(models.Ativo.nome).all()

    output = io.BytesIO()
    doc    = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=14, textColor=colors.HexColor("#1E3A34"))
    sub_style   = ParagraphStyle("s", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements = []
    titulo = "Relatório de Ativos"
    if status == "ativo":   titulo += " — Somente Ativos"
    elif status == "inativo": titulo += " — Somente Inativos"
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(f"Gerado em {_agora_br().strftime('%d/%m/%Y às %H:%M')}", sub_style))
    elements.append(Spacer(1, 0.4*cm))

    table_data = [["Nome", "Categoria", "Grupo", "Status", "Materiais em uso"]]
    for a in ativos:
        itens_ativos = sum(1 for i in a.itens if i.devolvido_em is None)
        table_data.append([
            a.nome,
            a.grupo.categoria.nome if a.grupo and a.grupo.categoria else "—",
            a.grupo.nome if a.grupo else "—",
            "Ativo" if a.ativo else "Inativo",
            str(itens_ativos),
        ])

    cw = [5.5*cm, 3.5*cm, 3.5*cm, 2*cm, 3*cm]
    table = Table(table_data, colWidths=cw, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E3A34")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(
        f"Total: {len(ativos)} registro(s) · © Todos os direitos reservados – github.com/Wbad-02",
        sub_style,
    ))
    doc.build(elements)
    output.seek(0)

    filename = f"ativos_{_agora_br().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Relatório de Notificações ─────────────────────────────────

@router.get("/notificacoes/excel")
def exportar_notificacoes_excel(
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(requer_admin),
):
    emails = db.query(models.NotificacaoEmail).filter(
        models.NotificacaoEmail.ativo == True
    ).order_by(models.NotificacaoEmail.tipo, models.NotificacaoEmail.email).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notificacoes"
    header_fill = PatternFill("solid", fgColor="1E3A34")
    header_font = Font(bold=True, color="FFFFFF")
    tipo_fills = {
        "retirada": PatternFill("solid", fgColor="E3F2FD"),
        "entrada":  PatternFill("solid", fgColor="E8F5E9"),
        "alerta":   PatternFill("solid", fgColor="FFF3E0"),
    }

    headers    = ["Tipo", "E-mail", "Intervalo (dias)", "Cadastrado em"]
    col_widths = [14, 36, 18, 18]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    tipo_label = {"retirada": "Retirada", "entrada": "Entrada", "alerta": "Alerta"}
    for ri, e in enumerate(emails, 2):
        row_data = [
            tipo_label.get(e.tipo, e.tipo),
            e.email,
            e.intervalo_dias if e.intervalo_dias else "—",
            e.criado_em.strftime("%d/%m/%Y"),
        ]
        fill = tipo_fills.get(e.tipo)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center" if ci != 2 else "left")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    output = io.BytesIO(); wb.save(output); output.seek(0)
    filename = f"notificacoes_{_agora_br().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Relatório de Entradas por NF-e ───────────────────────────

@router.get("/entradas-nfe")
def relatorio_entradas_nfe(
    mes: int,
    ano: int,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(requer_editor_ou_admin),
):
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func

    movs = (
        db.query(models.Movimentacao)
        .options(
            joinedload(models.Movimentacao.material).joinedload(models.Material.grupo).joinedload(models.GrupoMaterial.categoria),
        )
        .filter(
            models.Movimentacao.tipo == "entrada",
            models.Movimentacao.nf_numero.isnot(None),
            func.strftime("%m", models.Movimentacao.criado_em) == f"{mes:02d}",
            func.strftime("%Y", models.Movimentacao.criado_em) == str(ano),
        )
        .order_by(models.Movimentacao.criado_em.asc())
        .all()
    )

    resultado = []
    for m in movs:
        mat = m.material
        grupo = mat.grupo if mat else None
        categoria = grupo.categoria if grupo else None
        quantidade = m.quantidade or 0.0
        valor_unit = m.valor_unitario or 0.0
        resultado.append({
            "nf_numero":      m.nf_numero,
            "material_nome":  mat.nome if mat else "",
            "categoria_nome": categoria.nome if categoria else "",
            "grupo_nome":     grupo.nome if grupo else "",
            "quantidade":     quantidade,
            "unidade":        mat.unidade if mat else "",
            "valor_unitario": valor_unit,
            "subtotal":       round(quantidade * valor_unit, 2),
            "criado_em":      m.criado_em.isoformat(),
        })

    return resultado


# ── Consumo médio + previsão de ruptura ──────────────────────

@router.get("/consumo-medio")
def consumo_medio(
    meses: int = 3,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    from datetime import timedelta
    from sqlalchemy import func

    if not (1 <= meses <= 24):
        meses = 3

    corte = _agora_br() - timedelta(days=30 * meses)

    saidas_q = (
        db.query(
            models.Movimentacao.material_id,
            func.sum(models.Movimentacao.quantidade).label("total_saida"),
        )
        .filter(
            models.Movimentacao.tipo == "saida",
            models.Movimentacao.criado_em >= corte,
        )
        .group_by(models.Movimentacao.material_id)
        .all()
    )
    saidas_map = {r.material_id: float(r.total_saida or 0) for r in saidas_q}

    mats = (
        db.query(models.Material)
        .filter(models.Material.ativo == True)
        .join(models.GrupoMaterial)
        .order_by(models.GrupoMaterial.nome, models.Material.nome)
        .all()
    )

    resultado = []
    for m in mats:
        total = saidas_map.get(m.id, 0.0)
        media_mensal = round(total / meses, 2)
        if media_mensal > 0:
            dias = round((float(m.quantidade) / media_mensal) * 30)
        else:
            dias = None
        resultado.append({
            "material_id":          m.id,
            "material_nome":        m.nome,
            "categoria_nome":       m.grupo.categoria.nome if m.grupo and m.grupo.categoria else "",
            "grupo_nome":           m.grupo.nome if m.grupo else "",
            "estoque_atual":        float(m.quantidade),
            "unidade":              m.unidade,
            "consumo_total_periodo": total,
            "media_mensal":         media_mensal,
            "dias_cobertura":       dias,
        })

    resultado.sort(key=lambda x: (x["dias_cobertura"] is None, x["dias_cobertura"] or 0))
    return resultado


# ── Solicitações por material ─────────────────────────────────

@router.get("/solicitacoes-por-material")
def solicitacoes_por_material(
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(get_usuario_atual),
):
    from sqlalchemy import func, case

    rows = (
        db.query(
            models.SolicitacaoEstoque.material_id,
            func.count(models.SolicitacaoEstoque.id).label("total"),
            func.sum(case(
                (models.SolicitacaoEstoque.status == models.StatusSolicitacao.aprovado, 1), else_=0
            )).label("aprovados"),
            func.sum(case(
                (models.SolicitacaoEstoque.status == models.StatusSolicitacao.rejeitado, 1), else_=0
            )).label("rejeitados"),
            func.sum(case(
                (models.SolicitacaoEstoque.status == models.StatusSolicitacao.aguardando, 1), else_=0
            )).label("aguardando"),
        )
        .group_by(models.SolicitacaoEstoque.material_id)
        .order_by(func.count(models.SolicitacaoEstoque.id).desc())
        .all()
    )

    mat_ids = [r.material_id for r in rows]
    mats = {
        m.id: m
        for m in db.query(models.Material).filter(models.Material.id.in_(mat_ids)).all()
    } if mat_ids else {}

    resultado = []
    for r in rows:
        mat = mats.get(r.material_id)
        total = int(r.total or 0)
        aprovados = int(r.aprovados or 0)
        taxa = round((aprovados / total) * 100, 1) if total > 0 else 0.0
        resultado.append({
            "material_id":    r.material_id,
            "material_nome":  mat.nome if mat else "—",
            "categoria_nome": mat.grupo.categoria.nome if mat and mat.grupo and mat.grupo.categoria else "",
            "grupo_nome":     mat.grupo.nome if mat and mat.grupo else "",
            "total":          total,
            "aprovados":      aprovados,
            "rejeitados":     int(r.rejeitados or 0),
            "aguardando":     int(r.aguardando or 0),
            "taxa_aprovacao": taxa,
        })
    return resultado
