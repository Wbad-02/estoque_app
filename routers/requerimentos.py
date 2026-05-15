# © Todos os direitos reservados – github.com/Wbad-02
import io
import os
import re
import unicodedata

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

import models
import schemas
from auth import get_usuario_atual, requer_editor_ou_admin, registrar_log
from database import get_db
from email_service import disparar_notificacao

router = APIRouter(prefix="/api/requerimentos", tags=["requerimentos"])

_origins = os.environ.get("CORS_ORIGINS", "http://localhost:8000")
_URL_BASE = _origins.split(",")[0].strip().rstrip("/")


def _requer_criador_req(
    db:    Session        = Depends(get_db),
    atual: models.Usuario = Depends(get_usuario_atual),
) -> models.Usuario:
    """Admin+ passa sempre. Editor só passa se o seu e-mail está cadastrado em 'requerimento'."""
    if atual.grupo in (models.GrupoPermissao.admin, models.GrupoPermissao.mestre):
        return atual
    if atual.grupo == models.GrupoPermissao.editor:
        existe = db.query(models.NotificacaoEmail).filter(
            models.NotificacaoEmail.tipo  == "requerimento",
            models.NotificacaoEmail.ativo == True,
            models.NotificacaoEmail.email == atual.email,
        ).first()
        if existe:
            return atual
    raise HTTPException(403, "Acesso restrito: seu e-mail não está autorizado a criar requerimentos")


def _requer_aprovador_req(
    db:    Session        = Depends(get_db),
    atual: models.Usuario = Depends(get_usuario_atual),
) -> models.Usuario:
    """Admin+ passa sempre. Editor só passa se o seu e-mail está em 'requerimento_decisao'."""
    if atual.grupo in (models.GrupoPermissao.admin, models.GrupoPermissao.mestre):
        return atual
    if atual.grupo == models.GrupoPermissao.editor:
        existe = db.query(models.NotificacaoEmail).filter(
            models.NotificacaoEmail.tipo  == "requerimento_decisao",
            models.NotificacaoEmail.ativo == True,
            models.NotificacaoEmail.email == atual.email,
        ).first()
        if existe:
            return atual
    raise HTTPException(403, "Acesso restrito: seu e-mail não está autorizado a aprovar/rejeitar requerimentos")


def _slugify(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^\w\s-]", "", texto).strip().lower()
    texto = re.sub(r"[\s_-]+", "_", texto)
    return texto[:60]


def _build_out(req: models.Requerimento) -> schemas.RequerimentoOut:
    total = sum((i.quantidade or 1.0) * i.valor for i in req.itens)
    return schemas.RequerimentoOut(
        id=req.id,
        titulo=req.titulo,
        status=req.status.value if req.status else "aguardando",
        criado_em=req.criado_em,
        atualizado_em=req.atualizado_em,
        observacao=req.observacao,
        criador_nome=req.criador.nome if req.criador else "",
        aprovador_nome=req.aprovador.nome if req.aprovador else "",
        total=total,
        itens=[
            schemas.ItemRequerimentoOut(id=i.id, nome=i.nome, quantidade=i.quantidade or 1.0, valor=i.valor)
            for i in req.itens
        ],
    )


def _load(req_id: int, db: Session) -> models.Requerimento:
    req = (
        db.query(models.Requerimento)
        .options(
            joinedload(models.Requerimento.criador),
            joinedload(models.Requerimento.aprovador),
            joinedload(models.Requerimento.itens),
        )
        .filter(models.Requerimento.id == req_id)
        .first()
    )
    if not req:
        raise HTTPException(404, "Requerimento nao encontrado")
    return req


# ── POST / — criar ─────────────────────────────────────────────────────────────
@router.post("/", response_model=schemas.RequerimentoOut, status_code=201)
def criar_requerimento(
    payload: schemas.RequerimentoCreate,
    db: Session = Depends(get_db),
    atual: models.Usuario = Depends(_requer_criador_req),
):
    if not payload.itens:
        raise HTTPException(422, "O requerimento deve ter ao menos um item")

    req = models.Requerimento(
        titulo=payload.titulo.strip(),
        status=models.StatusRequerimento.aguardando,
        criado_por=atual.id,
    )
    db.add(req)
    db.flush()

    for item in payload.itens:
        db.add(models.ItemRequerimento(
            requerimento_id=req.id,
            nome=item.nome.strip(),
            quantidade=item.quantidade,
            valor=item.valor,
        ))

    db.commit()
    db.refresh(req)

    req = _load(req.id, db)
    total = sum(i.valor for i in req.itens)

    registrar_log(db, atual.id, "criar", "requerimento", req.id, payload.titulo)

    disparar_notificacao(db, "requerimento", {
        "titulo":       req.titulo,
        "total":        f"R$ {total:,.2f}",
        "itens_count":  str(len(req.itens)),
        "criador":      atual.nome,
        "link":         f"{_URL_BASE}/#requerimentos",
    })

    return _build_out(req)


# ── GET / — listar ─────────────────────────────────────────────────────────────
@router.get("/", response_model=list[schemas.RequerimentoOut])
def listar_requerimentos(
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(requer_editor_ou_admin),
):
    reqs = (
        db.query(models.Requerimento)
        .options(
            joinedload(models.Requerimento.criador),
            joinedload(models.Requerimento.aprovador),
            joinedload(models.Requerimento.itens),
        )
        .order_by(models.Requerimento.criado_em.desc())
        .all()
    )
    return [_build_out(r) for r in reqs]


# ── GET /{id} — detalhe ────────────────────────────────────────────────────────
@router.get("/{req_id}", response_model=schemas.RequerimentoOut)
def obter_requerimento(
    req_id: int,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(requer_editor_ou_admin),
):
    return _build_out(_load(req_id, db))


# ── POST /{id}/aprovar ─────────────────────────────────────────────────────────
@router.post("/{req_id}/aprovar", response_model=schemas.RequerimentoOut)
def aprovar_requerimento(
    req_id: int,
    body: schemas.AprovarRequerimentoBody = schemas.AprovarRequerimentoBody(),
    db: Session = Depends(get_db),
    atual: models.Usuario = Depends(_requer_aprovador_req),
):
    req = _load(req_id, db)
    if req.status != models.StatusRequerimento.aguardando:
        raise HTTPException(409, f"Requerimento ja esta '{req.status.value}'")

    req.status      = models.StatusRequerimento.aprovado
    req.aprovado_por = atual.id
    req.observacao   = body.observacao
    db.commit()
    db.refresh(req)

    req = _load(req_id, db)
    registrar_log(db, atual.id, "aprovar", "requerimento", req_id)

    criador_email = req.criador.email if req.criador else None
    disparar_notificacao(db, "requerimento_decisao", {
        "titulo":     req.titulo,
        "status":     "aprovado",
        "observacao": body.observacao or "—",
        "aprovador":  atual.nome,
    }, extras=[criador_email] if criador_email else None)

    return _build_out(req)


# ── POST /{id}/rejeitar ────────────────────────────────────────────────────────
@router.post("/{req_id}/rejeitar", response_model=schemas.RequerimentoOut)
def rejeitar_requerimento(
    req_id: int,
    body: schemas.RejeitarRequerimentoBody,
    db: Session = Depends(get_db),
    atual: models.Usuario = Depends(_requer_aprovador_req),
):
    req = _load(req_id, db)
    if req.status != models.StatusRequerimento.aguardando:
        raise HTTPException(409, f"Requerimento ja esta '{req.status.value}'")

    req.status      = models.StatusRequerimento.rejeitado
    req.aprovado_por = atual.id
    req.observacao   = body.observacao
    db.commit()
    db.refresh(req)

    req = _load(req_id, db)
    registrar_log(db, atual.id, "rejeitar", "requerimento", req_id, body.observacao)

    criador_email = req.criador.email if req.criador else None
    disparar_notificacao(db, "requerimento_decisao", {
        "titulo":     req.titulo,
        "status":     "rejeitado",
        "observacao": body.observacao,
        "aprovador":  atual.nome,
    }, extras=[criador_email] if criador_email else None)

    return _build_out(req)


# ── GET /{id}/excel ────────────────────────────────────────────────────────────
@router.get("/{req_id}/excel")
def exportar_excel(
    req_id: int,
    db: Session = Depends(get_db),
    _: models.Usuario = Depends(requer_editor_ou_admin),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    req = _load(req_id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Requerimento"

    # Paleta de cores
    COR_VERDE   = "1B3A2D"   # titulo
    COR_DOURADO = "C9A84C"   # cabecalho
    COR_BRANCO  = "FFFFFF"
    COR_TOTAL   = "F2F2F2"

    fill_verde   = PatternFill("solid", fgColor=COR_VERDE)
    fill_dourado = PatternFill("solid", fgColor=COR_DOURADO)
    fill_total   = PatternFill("solid", fgColor=COR_TOTAL)

    font_titulo  = Font(bold=True, color=COR_BRANCO, size=13)
    font_header  = Font(bold=True, color="000000", size=11)
    font_total   = Font(bold=True, size=11)
    font_normal  = Font(size=11)

    # ── Linha 1: titulo (A1:D1 merged) ────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = req.titulo
    ws["A1"].font      = font_titulo
    ws["A1"].fill      = fill_verde
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Linha 2: cabecalho ─────────────────────────────────────────────────────
    cabecalhos = {"A2": "Nome", "B2": "Quantidade", "C2": "Valor Unit. (R$)", "D2": "Subtotal (R$)"}
    for cel, texto in cabecalhos.items():
        ws[cel] = texto
        ws[cel].font      = font_header
        ws[cel].fill      = fill_dourado
        ws[cel].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Linhas dos itens ───────────────────────────────────────────────────────
    for linha, item in enumerate(req.itens, start=3):
        qtd      = item.quantidade or 1.0
        subtotal = qtd * item.valor
        ws.cell(row=linha, column=1, value=item.nome).font = font_normal
        c_qtd = ws.cell(row=linha, column=2, value=qtd)
        c_qtd.font = font_normal; c_qtd.number_format = "#,##0.##"; c_qtd.alignment = Alignment(horizontal="right")
        c_val = ws.cell(row=linha, column=3, value=item.valor)
        c_val.font = font_normal; c_val.number_format = "#,##0.00"; c_val.alignment = Alignment(horizontal="right")
        c_sub = ws.cell(row=linha, column=4, value=subtotal)
        c_sub.font = font_normal; c_sub.number_format = "#,##0.00"; c_sub.alignment = Alignment(horizontal="right")

    # ── Linha de total ─────────────────────────────────────────────────────────
    linha_total = 3 + len(req.itens)
    total = sum((i.quantidade or 1.0) * i.valor for i in req.itens)

    ws.merge_cells(f"A{linha_total}:C{linha_total}")
    cel_label = ws.cell(row=linha_total, column=1, value="TOTAL")
    cel_label.font = font_total; cel_label.fill = fill_total; cel_label.alignment = Alignment(horizontal="right")

    cel_total = ws.cell(row=linha_total, column=4, value=total)
    cel_total.font = font_total; cel_total.fill = fill_total
    cel_total.number_format = "#,##0.00"; cel_total.alignment = Alignment(horizontal="right")

    # Larguras das colunas
    ws.column_dimensions[get_column_letter(1)].width = 44
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 18
    ws.column_dimensions[get_column_letter(4)].width = 18

    # Serializar em memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    slug     = _slugify(req.titulo)
    filename = f"requerimento_{req_id}_{slug}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── GET /modelo-excel — template em branco ────────────────────────────────────
@router.get("/modelo-excel")
def baixar_modelo_excel(
    _: models.Usuario = Depends(requer_editor_ou_admin),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Requerimento"

    fill_verde   = PatternFill("solid", fgColor="1B3A2D")
    fill_dourado = PatternFill("solid", fgColor="C9A84C")

    ws.merge_cells("A1:D1")
    ws["A1"] = "MODELO — Requerimento de Compra"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = fill_verde
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    cabecalhos = ["Nome do Item", "Quantidade", "Valor Unitário (R$)", "Subtotal (R$)"]
    for col, texto in enumerate(cabecalhos, start=1):
        c = ws.cell(row=2, column=col, value=texto)
        c.font      = Font(bold=True, size=11)
        c.fill      = fill_dourado
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Três linhas de exemplo
    exemplos = [
        ("Caneta esferográfica azul", 10, 2.50),
        ("Resma de papel A4", 5, 24.90),
        ("Pasta arquivo AZ", 3, 18.00),
    ]
    for linha, (nome, qtd, val) in enumerate(exemplos, start=3):
        ws.cell(row=linha, column=1, value=nome)
        c_qtd = ws.cell(row=linha, column=2, value=qtd)
        c_qtd.number_format = "#,##0.##"; c_qtd.alignment = Alignment(horizontal="right")
        c_val = ws.cell(row=linha, column=3, value=val)
        c_val.number_format = "#,##0.00"; c_val.alignment = Alignment(horizontal="right")
        c_sub = ws.cell(row=linha, column=4, value=f"=B{linha}*C{linha}")
        c_sub.number_format = "#,##0.00"; c_sub.alignment = Alignment(horizontal="right")

    ws.column_dimensions[get_column_letter(1)].width = 44
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(4)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_requerimento.xlsx"'},
    )


# ── POST /importar-excel — cria requerimento a partir de planilha ─────────────
@router.post("/importar-excel", response_model=schemas.RequerimentoOut, status_code=201)
def importar_excel(
    titulo: str,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    atual: models.Usuario = Depends(_requer_criador_req),
):
    from openpyxl import load_workbook

    if not arquivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(422, "Envie um arquivo .xlsx")

    try:
        conteudo = arquivo.file.read()
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(422, "Arquivo Excel inválido ou corrompido")

    itens_raw = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        nome = str(row[0]).strip() if row[0] not in (None, "") else ""
        if not nome or nome.upper() in ("TOTAL", "NOME DO ITEM"):
            continue
        try:
            qtd = float(row[1]) if row[1] not in (None, "") else 1.0
            val = float(row[2]) if row[2] not in (None, "") else 0.0
        except (TypeError, ValueError):
            continue
        if qtd <= 0 or val <= 0:
            continue
        itens_raw.append({"nome": nome, "quantidade": qtd, "valor": val})

    if not itens_raw:
        raise HTTPException(422, "Nenhum item válido encontrado na planilha (verifique o modelo)")

    titulo = titulo.strip()
    if not titulo:
        raise HTTPException(422, "Informe o título do requerimento")

    req = models.Requerimento(
        titulo=titulo,
        status=models.StatusRequerimento.aguardando,
        criado_por=atual.id,
    )
    db.add(req)
    db.flush()

    for item in itens_raw:
        db.add(models.ItemRequerimento(
            requerimento_id=req.id,
            nome=item["nome"],
            quantidade=item["quantidade"],
            valor=item["valor"],
        ))

    db.commit()
    db.refresh(req)
    req = _load(req.id, db)
    total = sum((i.quantidade or 1.0) * i.valor for i in req.itens)

    registrar_log(db, atual.id, "criar", "requerimento", req.id, titulo)

    disparar_notificacao(db, "requerimento", {
        "titulo":      req.titulo,
        "total":       f"R$ {total:,.2f}",
        "itens_count": str(len(req.itens)),
        "criador":     atual.nome,
        "link":        f"{_URL_BASE}/#requerimentos",
    })

    return _build_out(req)
